from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Buat workbook dan sheet
wb = Workbook()
ws = wb.active
ws.title = "Harian"

headers = [
    "Tanggal", "Signature 60 mnt", "Signature 90 mnt", "Signature 120 mnt",
    "Atletic 90 mnt", "Atletic 120 mnt", "Total Member", "Total Komisi (Rp)"
]
ws.append(headers)

# Gaya
header_fill = PatternFill(start_color="9DC183", end_color="9DC183", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
center_align = Alignment(horizontal="center", vertical="center")
border_style = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))

for col, header in enumerate(headers, 1):
    c = ws.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center_align
    c.border = border_style
    ws.column_dimensions[get_column_letter(col)].width = 20

# Rumus otomatis
for i in range(2, 33):
    ws[f"H{i}"] = f"=(B{i}*23500)+(C{i}*35250)+(D{i}*47000)+(E{i}*42250)+(F{i}*63000)+(G{i}*5000)"

# Total bulanan
total_row = 34
ws[f"A{total_row}"] = "TOTAL BULANAN"
ws[f"A{total_row}"].font = Font(bold=True, color="004D00")
ws[f"H{total_row}"] = "=SUM(H2:H33)"
ws[f"H{total_row}"].font = Font(bold=True, color="004D00")

# Catatan
notes_start = total_row + 3
ws[f"A{notes_start}"] = "Keterangan:"
ws[f"A{notes_start}"].font = Font(bold=True, color="004D00")

notes = [
    "Signature Massage: 60mnt=23.500 | 90mnt=35.250 | 120mnt=47.000",
    "Atletic Massage: 90mnt=42.250 | 120mnt=63.000",
    "Bonus Member: Rp5.000 per customer member",
    "Isi tanggal dan jumlah customer setiap hari. Total komisi otomatis dihitung.",
    "Gunakan sheet 'Rekap Bulanan' untuk laporan gaji tiap bulan (gajian tiap 27)."
]
for i, note in enumerate(notes, start=notes_start + 1):
    ws[f"A{i}"] = f"• {note}"

# Sheet 2
rekap = wb.create_sheet("Rekap Bulanan")
rekap_headers = ["Bulan", "Total Hari Kerja", "Total Customer", "Total Member", "Total Komisi (Rp)", "Gajian Tanggal 27"]
rekap.append(rekap_headers)
for col, header in enumerate(rekap_headers, 1):
    c = rekap.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center_align
    c.border = border_style
    rekap.column_dimensions[get_column_letter(col)].width = 25

rekap.cell(row=2, column=1, value="Template (isi manual tiap bulan)")

# Simpan dengan nama file per bulan
bulan = datetime.now().strftime("%B_%Y")
filename = f"Komisi_Kokuo_{bulan}.xlsx"
wb.save(filename)
print(f"✅ File berhasil dibuat: {filename}")
